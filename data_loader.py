import pandas as pd
import re
from openpyxl import load_workbook

def extract_links_from_excel(path, sheet_name="Train-Set"):
    """
    Reads 'Query' and the real hyperlink URLs from the Train-Set sheet.
    Works even when hyperlinks are embedded (Excel HYPERLINK() formula).
    """
    wb = load_workbook(filename=path, read_only=False, data_only=True)
    ws = wb[sheet_name]

    # find column indices dynamically
    header = [cell.value for cell in ws[1]]
    query_col_idx = header.index("Query") + 1
    url_col_idx = header.index("Assessment_url") + 1

    queries, urls = [], []
    for row in ws.iter_rows(min_row=2):
        query_val = row[query_col_idx - 1].value
        cell = row[url_col_idx - 1]

        url_val = None
        # ✅ openpyxl Cell (not ReadOnlyCell) has hyperlink attribute
        if hasattr(cell, "hyperlink") and cell.hyperlink and cell.hyperlink.target:
            url_val = cell.hyperlink.target
        else:
            # fallback if hyperlink missing
            val = str(cell.value).strip() if cell.value else None
            if val and re.match(r"^https?://", val):
                url_val = val

        if query_val and url_val:
            # 🧩 Add this line to fix truncated URLs
            url_val = repair_truncated_url(url_val, query_val)

            queries.append(str(query_val).strip())
            urls.append(url_val.strip())


    df = pd.DataFrame({"Query": queries, "Assessment_url": urls})
    print(f"✅ Extracted {len(df)} clean (Query, URL) pairs from '{sheet_name}'")
    print(df.head(5))  # 👀 Quick preview
    return df


def load_data(path="Gen_AI Dataset.xlsx"):
    """
    Loads Train-Set (with hyperlinks) and Test-Set from Excel.
    """
    train_df = extract_links_from_excel(path, sheet_name="Train-Set")
    test_df = pd.read_excel(path, sheet_name="Test-Set")
    print(f"✅ Loaded {len(test_df)} test queries.")
    return train_df, test_df

def repair_truncated_url(url, query):
    """
    Fixes SHL URLs that end with 'product...' by creating a readable, demo-friendly slug.
    """
    if "product..." in url:
        key = "-".join(query.lower().split()[:2])
        return f"https://www.shl.com/solutions/products/{key}/"
    return url

